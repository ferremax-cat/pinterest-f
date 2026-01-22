import openpyxl
import json
import os
from pathlib import Path

def leer_funcionalidades_maestro():
    """Lee el archivo funcionalidades_maestro.xlsx y retorna diccionario"""
    
    # Rutas relativas desde la carpeta scripts
    excel_path = Path(__file__).parent.parent / 'excel' / 'funcionalidades_maestro.xlsx'
    
    print(f"📂 Leyendo: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo {excel_path}")
        return None
    
    resultado = {
        "funcionalidades": {},
        "roles": {}
    }
    
    # --- Leer hoja "Funcionalidades" ---
    try:
        sheet_func = wb["Funcionalidades"]
    except KeyError:
        print("❌ ERROR: No se encontró la hoja 'Funcionalidades'")
        return None
    
    # Saltar la primera fila (encabezados)
    for row in sheet_func.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Si hay ID
            func_id = str(row[0]).strip()
            nombre = str(row[1]).strip() if row[1] else ""
            descripcion = str(row[2]).strip() if row[2] else ""
            icono = str(row[3]).strip() if row[3] else ""
            activa = str(row[4]).strip().upper() == "TRUE" if row[4] else True
            
            resultado["funcionalidades"][func_id] = {
                "id": func_id,
                "nombre": nombre,
                "descripcion": descripcion,
                "icono": icono,
                "activa": activa
            }
    
    print(f"✅ Leídas {len(resultado['funcionalidades'])} funcionalidades")
    
    # --- Leer hoja "Roles" ---
    try:
        sheet_roles = wb["Roles"]
    except KeyError:
        print("❌ ERROR: No se encontró la hoja 'Roles'")
        return None
    
    for row in sheet_roles.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Si hay rol
            rol_id = str(row[0]).strip()
            nombre_rol = str(row[1]).strip() if row[1] else ""
            funcionalidades_str = str(row[2]).strip() if row[2] else ""
            
            # Convertir string separado por comas en lista
            if funcionalidades_str:
                funcionalidades_list = [f.strip() for f in funcionalidades_str.split(",")]
            else:
                funcionalidades_list = []
            
            resultado["roles"][rol_id] = {
                "nombre": nombre_rol,
                "funcionalidades": funcionalidades_list
            }
    
    print(f"✅ Leídos {len(resultado['roles'])} roles")
    
    wb.close()
    return resultado


def leer_usuarios_funcionalidades():
    """Lee el archivo usuarios_funcionalidades.xlsx y retorna diccionario"""
    
    excel_path = Path(__file__).parent.parent / 'excel' / 'usuarios_funcionalidades.xlsx'
    
    print(f"📂 Leyendo: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo {excel_path}")
        return None
    
    resultado = {
        "usuarios": {}
    }
    
    # --- Leer hoja "Usuarios" ---
    try:
        sheet_usuarios = wb["Usuarios"]
    except KeyError:
        print("❌ ERROR: No se encontró la hoja 'Usuarios'")
        return None
    
    for row in sheet_usuarios.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Si hay clave
            clave = str(row[0]).strip()
            nombre = str(row[1]).strip() if row[1] else ""
            numero_cuenta = str(row[2]).strip() if row[2] else ""
            rol = str(row[3]).strip() if row[3] else ""
            
            usuario_data = {
                "clave": clave,
                "nombre": nombre,
                "rol": rol
            }
            
            # Agregar numero_cuenta solo si tiene valor
            if numero_cuenta:
                usuario_data["numero_cuenta"] = numero_cuenta
            
            # Funcionalidades extra (columna 5)
            if row[4]:
                extras_str = str(row[4]).strip()
                usuario_data["funcionalidades_extra"] = [f.strip() for f in extras_str.split(",")]
            
            # Funcionalidades bloqueadas (columna 6)
            if row[5]:
                bloqueadas_str = str(row[5]).strip()
                usuario_data["funcionalidades_bloqueadas"] = [f.strip() for f in bloqueadas_str.split(",")]
            
            resultado["usuarios"][clave] = usuario_data
    
    print(f"✅ Leídos {len(resultado['usuarios'])} usuarios")
    
    wb.close()
    return resultado


def guardar_json(data, filename):
    """Guarda diccionario como JSON en la carpeta json/"""
    
    json_path = Path(__file__).parent.parent / 'json' / filename
    
    # Crear carpeta json si no existe
    json_path.parent.mkdir(exist_ok=True)
    
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"✅ Guardado: {json_path}")
        return True
    except Exception as e:
        print(f"❌ ERROR al guardar {filename}: {e}")
        return False


def main():
    """Función principal"""
    
    print("\n" + "="*60)
    print("🔄 CONVERTIDOR EXCEL → JSON")
    print("   Usuarios y Funcionalidades")
    print("="*60 + "\n")
    
    # Leer funcionalidades_maestro.xlsx
    print("📋 PASO 1: Leyendo funcionalidades maestro...")
    funcionalidades_data = leer_funcionalidades_maestro()
    
    if not funcionalidades_data:
        print("\n❌ Error al leer funcionalidades_maestro.xlsx")
        return
    
    # Leer usuarios_funcionalidades.xlsx
    print("\n📋 PASO 2: Leyendo usuarios y funcionalidades...")
    usuarios_data = leer_usuarios_funcionalidades()
    
    if not usuarios_data:
        print("\n❌ Error al leer usuarios_funcionalidades.xlsx")
        return
    
    # Guardar JSONs
    print("\n💾 PASO 3: Guardando archivos JSON...")
    
    exito1 = guardar_json(funcionalidades_data, 'funcionalidades.json')
    exito2 = guardar_json(usuarios_data, 'funcionalidades_usuarios.json')
    
    # Resumen final
    print("\n" + "="*60)
    if exito1 and exito2:
        print("✅ CONVERSIÓN COMPLETADA EXITOSAMENTE")
        print(f"   • {len(funcionalidades_data['funcionalidades'])} funcionalidades")
        print(f"   • {len(funcionalidades_data['roles'])} roles")
        print(f"   • {len(usuarios_data['usuarios'])} usuarios")
    else:
        print("⚠️  CONVERSIÓN COMPLETADA CON ERRORES")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()